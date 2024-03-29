from email import utils
import re, os
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.contrib.admin.views.decorators import staff_member_required

from .models import Section, Exam, SectionInstance, ExamInstance, Student
from questions.models import Question, Answer, Result, Student_Answer
from .utils.mappings import act_math_category_map, sat_math_category_map, sat_english_category_map, act_english_category_map
from .utils.helpers import getMaterialUrlFromQuestion
# Create your views here.

@staff_member_required
def file_upload(request):
    if request.method == 'POST':

        questions_file = request.FILES['questions_file']

        #TODO: CANNOT ASSUME THAT THE FILES BEING UPLOADED ARE IN ORDER
        material_files_index = 0
        if not questions_file.name.endswith('.xls') and not questions_file.name.endswith('.xlsx'):
            return render(request, 'exams/file_upload.html', {'error': 'You did not upload a valid Excel file.'})

        wb = load_workbook(filename=request.FILES['questions_file'].file)
        worksheet = wb["Sheet1"]

        #processing variables
        exam_name = ''
        current_section = ''
        question_number = 1
        num_passages = 0

        #TODO FIX THE SAME QUESTION TEXT ERROR
        no_question_index = 1
        #Declare object variables to reduce querying to database
        exam_object = None
        section_object = None
        question_object = None
        exam_type = None

        # Variable to hold the ordering of the sections (assumes that the excel file contains the sections in the right order)
        section_order = 1

        for row in worksheet.iter_rows(min_row=2):
            #checks if the 'section' column is empty - this is when processing should stop
            if row[2].value is None:
                break

            #create new exam
            #for cell in row:
                #print(cell.value)

            if question_object is None:
                exam_name = row[0].value
                if Exam.objects.filter(name=exam_name).count() > 0:
                    #Exam.objects.get(name=exam_name).delete()
                    return render(request, 'exams/file_upload.html', {'error': 'An exam with the same name is already in the database. Please rename this exam.'})

                #EXAM TYPE IS SAT OR ACT
                exam_type = row[1].value

                exam_object, created = Exam.objects.get_or_create(
                    name = exam_name,
                    type = row[1].value,
                )
                if exam_type is None or not (exam_type == 'ACT' or exam_type == 'SAT' or exam_type == 'DIAGNOSTIC'):
                    Exam.objects.get(name=exam_name).delete()
                    return render(request, 'exams/file_upload.html', {'error': 'The exam has been uploaded without an accepted exam type. Please specify the type in the Excel file (SAT or ACT)'})
            #update the previous section's num_passage field
            if section_object is not None:
                section_object.num_passages = num_passages
                section_object.save()
                num_passages = 0

            #Checks if the cell in the section column is empty
            if row[2].value is None:
                Exam.objects.get(name=exam_name).delete()
                return render(request, 'exams/file_upload.html', {'error': 'A question has a missing section, please specify the section in the Excel file.'})
            #if we have a new section, then we create a new section
            if current_section != row[2].value.lower():
                current_section = row[2].value.lower()

                num_questions = 0
                time = 0
                #resets the question_number to 1 for a new section
                question_number = 1
                section_name = None

                if exam_type == 'SAT':
                    if current_section == 'reading':
                            num_questions = 52
                            time = 65
                            section_name = 'Reading'
                    elif current_section == 'writing':
                        num_questions = 44
                        time = 35
                        section_name = 'Writing and Language'
                    elif current_section == 'math1':
                        num_questions = 20
                        time = 25
                        section_name = 'Math (No Calculator)'
                    elif current_section =='math2':
                        num_questions = 38
                        time = 55
                        section_name = 'Math (Calculator)'
                elif exam_type == 'ACT':
                    # SECTION NAMES OF THE ACT
                    if current_section == 'english' or current_section == 'writing':
                        num_questions = 75
                        time = 45
                        section_name = 'English'
                    elif current_section == 'math':
                        num_questions = 60
                        time = 60
                        section_name = 'Math'
                    elif current_section == 'reading':
                        num_questions = 40
                        time = 35
                        section_name = 'Reading'
                    elif current_section == 'science':
                        num_questions = 40
                        time = 35
                        section_name = 'Science'
                elif exam_type == 'DIAGNOSTIC':
                    if current_section == 'writing':
                        num_questions = 22
                        time = 17
                        section_name = 'Writing and Language - SAT'
                    elif current_section == 'math1':
                        num_questions = 10
                        time = 10
                        section_name = 'Math (No Calculator) - SAT'
                    elif current_section == 'math2':
                        num_questions = 15
                        time = 21
                        section_name = 'Math (Calculator) - SAT'
                    elif current_section == 'math':
                        num_questions = 30
                        time = 30
                        section_name = 'Math - ACT'
                    elif current_section == 'reading':
                        num_questions = 21
                        time = 25
                        section_name = 'Reading - SAT'
                    elif current_section == 'english':
                        num_questions = 30
                        time = 18
                        section_name = 'English - ACT'
                    elif current_section == 'science':
                        num_questions = 21
                        time = 18
                        section_name = 'Science - ACT'

                section_object, created = Section.objects.get_or_create(
                    name = section_name,
                    type = 'english' if (current_section == 'writing' and exam_type == 'ACT') else current_section,
                    exam = exam_object,
                    num_questions = num_questions,
                    time = time,
                    ordering = section_order
                )

                section_order += 1
            #create questions
            question_text = row[4].value

            if question_text is not None:
                question_text = question_text.replace('"', '&quot;')
            else:
                question_text = "no question" + str(no_question_index)
                no_question_index += 1

            question_text = question_text.replace("\n", "\\n")

            # handling of non float values in the column
            #import pdb; pdb.set_trace()
            question_passage = int(row[3].value) if (isinstance(row[3].value, int) or isinstance(row[3].value, float)) else None
            #is_open_ended_math_question = True if isinstance(row[10].value, float) or isinstance(row[10].value, int) or (len(row[10].value) > 1) else False
            is_open_ended_math_question = True if not str(row[10].value).isalpha() else False
            correct_answer = row[10].value if isinstance(row[10].value, float) or isinstance(row[10].value, int) else row[10].value.upper()

            question_categories = row[11].value

            # replace the "RG" category to "GF" only for the science section
            if current_section == "science":
                for question_category in question_categories:
                    if "RG" in question_category:
                        question_category = "GF"

            if question_passage is not None:
                if question_passage > num_passages:
                    num_passages = question_passage
            question_object, created = Question.objects.get_or_create(
                question_number = question_number,
                text = question_text,
                section = section_object,
                passage = question_passage,
                correct_answer = correct_answer,
                exam = exam_object,
                categories = question_categories,
            )
            question_number += 1

            #check if multiple choice question has any empty values
            if not is_open_ended_math_question:
                if (row[5].value == None or row[6].value == None or row[7].value == None or row[8].value == None):
                    Exam.objects.get(name=exam_name).delete()

                    return render(request, 'exams/file_upload.html', {'error': f'Question {question_number} in section {section_object.name} has a missing answer'})

            #create answers to question
            is_abcd_question = False
            is_fghjk_question = False
            if correct_answer == 'F' or correct_answer == 'G' or correct_answer == 'H' or correct_answer == 'J' or correct_answer == 'K':
                is_fghjk_question = True
            elif correct_answer == 'A' or correct_answer == 'B' or correct_answer == 'C' or correct_answer == 'D' or correct_answer == 'E':
                is_abcd_question = True
            elif not is_open_ended_math_question:
                Exam.objects.get(name=exam_name).delete()
                return render(request, 'exams/file_upload.html', {'error': f'Question {question_number} in section {section_object.name} has an invalid answer key'})

            answer_object_A, created = Answer.objects.get_or_create(
                text = row[5].value,
                letter = 'F' if is_fghjk_question else 'A',
                question = question_object
            )
            answer_object_B, created = Answer.objects.get_or_create(
                text = row[6].value,
                letter = 'G' if is_fghjk_question else 'B',
                question = question_object
            )
            answer_object_C, created = Answer.objects.get_or_create(
                text = row[7].value,
                letter = 'H' if is_fghjk_question else 'C',
                question = question_object
            )
            answer_object_D, created = Answer.objects.get_or_create(
                text = row[8].value,
                letter = 'J' if is_fghjk_question else 'D',
                question = question_object
            )
            if row[9].value is not None:
                answer_object_E, created = Answer.objects.get_or_create(
                    text = row[9].value,
                    letter = 'K' if is_fghjk_question else 'E',
                    question = question_object
                )

            """
            if exam_type != 'DIAGNOSTIC':
                answer_object_A, created = Answer.objects.get_or_create(
                    text = row[5].value,
                    letter = 'F' if (question_number % 2 == 1 and exam_type == 'ACT') else 'A' ,
                    question = question_object
                )

                answer_object_B, created = Answer.objects.get_or_create(
                    text = row[6].value,
                    letter = 'G' if (question_number % 2 == 1 and exam_type == 'ACT') else 'B' ,
                    question = question_object
                )

                answer_object_C, created = Answer.objects.get_or_create(
                    text = row[7].value,
                    letter = 'H' if (question_number % 2 == 1 and exam_type == 'ACT') else 'C' ,
                    question = question_object
                )

                answer_object_D, created = Answer.objects.get_or_create(
                    text = row[8].value,
                    letter = 'J' if (question_number % 2 == 1 and exam_type == 'ACT') else 'D' ,
                    question = question_object
                )
            elif exam_type == 'DIAGNOSTIC':
                answer_object_A, created = Answer.objects.get_or_create(
                    text = row[5].value,
                    letter = 'F' if (question_number % 2 == 1 and (section_object.type == 'english' or section_object.type == 'math' or section_object.type == 'science')) else 'A' ,
                    question = question_object
                )

                answer_object_B, created = Answer.objects.get_or_create(
                    text = row[6].value,
                    letter = 'G' if (question_number % 2 == 1 and (section_object.type == 'english' or section_object.type == 'math' or section_object.type == 'science')) else 'B' ,
                    question = question_object
                )

                answer_object_C, created = Answer.objects.get_or_create(
                    text = row[7].value,
                    letter = 'H' if (question_number % 2 == 1 and (section_object.type == 'english' or section_object.type == 'math' or section_object.type == 'science')) else 'C' ,
                    question = question_object
                )

                answer_object_D, created = Answer.objects.get_or_create(
                    text = row[8].value,
                    letter = 'J' if (question_number % 2 == 1 and (section_object.type == 'english' or section_object.type == 'math' or section_object.type == 'science')) else 'D' ,
                    question = question_object
                )

            #Creates the fifth answer choice for the ACT exam
            if exam_type == 'ACT' or exam_type == 'DIAGNOSTIC':
                if row[9].value is not None:
                    answer_object_E, created = Answer.objects.get_or_create(
                        text = row[9].value,
                        letter = 'K' if (question_number % 2 == 1) else 'E',
                        question = question_object
                    )

        """
        #assign images to the respective questions in each section
        #NOTE: for reading passages, the image should be representative of the entire passage
        if exam_type == 'SAT':
            sat_reading_passages = request.FILES.getlist('sat_reading_passages')
            sat_writing_passages = request.FILES.getlist('sat_writing_passages')
            sat_nocalc_materials = request.FILES.getlist('sat_nocalc_materials')
            sat_calc_materials = request.FILES.getlist('sat_calc_materials')
            for file in sat_reading_passages:
                filename = file.name
                passage_num = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='reading', exam=exam_object)
                question_object = Question.objects.filter(section=section_object, passage=passage_num).order_by('question_number')[0]
                question_object.material = file
                question_object.save()

            for file in sat_writing_passages:
                filename = file.name
                passage_num = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='writing', exam=exam_object)
                question_object = Question.objects.filter(section=section_object, passage=passage_num).order_by('question_number')[0]
                question_object.material = file
                question_object.save()

            for file in sat_nocalc_materials:
                filename = file.name
                question_no = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='math1', exam=exam_object)
                question_object = Question.objects.get(question_number=question_no, section=section_object)
                question_object.material = file
                question_object.save()

            for file in sat_calc_materials:
                filename = file.name
                question_no = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='math2', exam=exam_object)
                question_object = Question.objects.get(question_number=question_no, section=section_object)
                question_object.material = file
                question_object.save()
        elif exam_type == 'ACT':
            act_english_passages = request.FILES.getlist('act_english_passages')
            act_math_materials = request.FILES.getlist('act_math_materials')
            act_reading_passages = request.FILES.getlist('act_reading_passages')
            act_science_passages = request.FILES.getlist('act_science_passages')
            for file in act_english_passages:
                filename = file.name
                passage_num = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='english', exam=exam_object)
                question_object = Question.objects.filter(section=section_object, passage=passage_num).order_by('question_number')[0]
                question_object.material = file
                question_object.save()

            for file in act_math_materials:
                filename = file.name
                question_no = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='math', exam=exam_object)
                question_object = Question.objects.get(question_number=question_no, section=section_object)
                question_object.material = file
                question_object.save()

            for file in act_reading_passages:
                filename = file.name
                passage_num = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='reading', exam=exam_object)
                question_object = Question.objects.filter(section=section_object, passage=passage_num).order_by('question_number')[0]
                question_object.material = file
                question_object.save()

            for file in act_science_passages:
                filename = file.name
                passage_num = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='science', exam=exam_object)
                question_object = Question.objects.filter(section=section_object, passage=passage_num).order_by('question_number')[0]
                question_object.material = file
                question_object.save()
        elif exam_type == 'DIAGNOSTIC':
            sat_reading_passages = request.FILES.getlist('sat_reading_passages')
            sat_writing_passages = request.FILES.getlist('sat_writing_passages')
            sat_nocalc_materials = request.FILES.getlist('sat_nocalc_materials')
            sat_calc_materials = request.FILES.getlist('sat_calc_materials')

            act_english_passages = request.FILES.getlist('act_english_passages')
            act_math_materials = request.FILES.getlist('act_math_materials')
            act_science_passages = request.FILES.getlist('act_science_passages')
            for file in act_english_passages:
                filename = file.name
                passage_num = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='english', exam=exam_object)
                question_object = Question.objects.filter(section=section_object, passage=passage_num).order_by('question_number')[0]
                question_object.material = file
                question_object.save()

            for file in act_math_materials:
                filename = file.name
                question_no = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='math', exam=exam_object)
                question_object = Question.objects.get(question_number=question_no, section=section_object)
                question_object.material = file
                question_object.save()

            for file in act_science_passages:
                filename = file.name
                passage_num = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='science', exam=exam_object)
                question_object = Question.objects.filter(section=section_object, passage=passage_num).order_by('question_number')[0]
                question_object.material = file
                question_object.save()

            for file in sat_reading_passages:
                filename = file.name
                passage_num = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='reading', exam=exam_object)
                question_object = Question.objects.filter(section=section_object, passage=passage_num).order_by('question_number')[0]
                question_object.material = file
                question_object.save()

            for file in sat_writing_passages:
                filename = file.name
                passage_num = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='writing', exam=exam_object)
                question_object = Question.objects.filter(section=section_object, passage=passage_num).order_by('question_number')[0]
                question_object.material = file
                question_object.save()

            for file in sat_nocalc_materials:
                filename = file.name
                question_no = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='math1', exam=exam_object)
                question_object = Question.objects.get(question_number=question_no, section=section_object)
                question_object.material = file
                question_object.save()

            for file in sat_calc_materials:
                filename = file.name
                question_no = int(re.findall(r'\d+', filename)[0])
                section_object = Section.objects.get(type='math2', exam=exam_object)
                question_object = Question.objects.get(question_number=question_no, section=section_object)
                question_object.material = file
                question_object.save()
        return render(request, 'exams/file_upload.html', {'success': 'Exam successfully uploaded to database.'})

    return render(request, 'exams/file_upload.html')

def index(request):
    print(request.user)
    #REDIRECTS TO THE EXAM LIST VIEW IF THE USER IS ALREADY LOGGED IN
    if request.user.is_authenticated:
        return redirect('exams:exam-list-view')
    return render(request, 'index.html', {})

@login_required
def exam_list_view(request):
    user = request.user
    exams = Exam.objects.all().order_by('name')
    # Calculate the number of sections completed for each exam
    num_sections = []
    sections_completed = []
    for exam in exams:
        num_sections.append(len(exam.get_sections()))
        sections_completed.append(Result.objects.filter(user=user, exam=exam).count())

    exam_info = zip(exams, sections_completed, num_sections)
    # Get the most recent exam
    recent_exam = None

    payment_status = None

    student:Student = Student.objects.get(user=user)
    if student.recent_exam is not None:
        recent_exam = student.recent_exam.name
    # Check if a payment was recently attempted
    if student.payment_status != 'NONE':
        payment_status = student.payment_status
        student.payment_status = 'NONE'
        student.save()

    context = {
        'exams':exams,
        'exam_info':exam_info,
        'sections_completed':sections_completed,
        'num_sections':num_sections,
        'recent_exam': recent_exam,
        'payment_status': payment_status,
        'student': student,
    }
    return render(request, 'exams/exam_list.html', context)

@login_required
def exam_list_data_view(request, pk):

    exam = Exam.objects.get(pk=pk)
    sections = exam.get_sections()
    user = request.user
    results = Result.objects.filter(exam=exam, user=user)

    data = []

    # Appending the raw scores
    """
    if results.exists() > 0:
        for result in results:
            raw_scores.append({result.section.type: result.score})
    """

    for section in sections:
        #import pdb; pdb.set_trace()
        minutes = None
        seconds = None
        raw_score = None
        scaled_score = None
        direct_section_url = None
        section_directions_url = reverse('exams:section-directions-view', kwargs={'pk':pk, 'section_name':section.type})
        if SectionInstance.objects.filter(user=user, exam=exam, section=section).exists():
            section_instance = SectionInstance.objects.get(user=user, exam=exam, section=section)
            minutes = section_instance.minutes_left
            seconds = section_instance.seconds_left
            has_section_instance = True
            direct_section_url = reverse('exams:section-view', kwargs={'pk':pk, 'section_name':section.type})
        if Result.objects.filter(user=user, exam=exam, section=section).exists():
            result = Result.objects.get(user=user, exam=exam, section=section)
            raw_score = result.raw_score
            scaled_score = result.scaled_score

        data.append({section.type: [raw_score, scaled_score, minutes, seconds, section_directions_url, direct_section_url]})

    # Get the is_extended_time boolean
    if ExamInstance.objects.filter(user=user, exam=exam).exists():
        exam_instance = ExamInstance.objects.get(user=user, exam=exam)
        is_extended_time = exam_instance.is_extended_time
    else:
        is_extended_time = False

    return JsonResponse({
        'data': data,
        'is_extended_time': is_extended_time,
    })

def exam_list_recent_exam_view(request):
    user = request.user

    # Get the most recent exam
    recent_exam = None
    recent_exam_type = None

    student = Student.objects.get(user=user)
    if student.recent_exam is not None:
        recent_exam = student.recent_exam.name
        recent_exam_type = student.recent_exam.type

    return JsonResponse({'recent_exam': recent_exam, 'recent_exam_type': recent_exam_type})
@login_required
def exam_list_reset_view(request, pk):
    exam = Exam.objects.get(pk=pk)
    user = request.user

    section_instances = SectionInstance.objects.filter(user=user, exam=exam)
    student_answers = Student_Answer.objects.filter(user=user, exam=exam)
    results = Result.objects.filter(user=user, exam=exam)
    exam_instance = ExamInstance.objects.filter(user=user, exam=exam)

    if section_instances.count() > 0:
        section_instances.delete()

    if student_answers.count() > 0:
        student_answers.delete()

    if results.count() > 0:
        results.delete()

    if exam_instance.count() > 0:
        exam_instance.delete()

    return JsonResponse({})

@login_required
def exam_list_change_time_view(request, pk):

    exam = Exam.objects.get(pk=pk)
    user = request.user
    is_extended_time = request.POST['is_extended_time']
    is_extended_time = True if is_extended_time == 'true' else False

    exam_instance, updated = ExamInstance.objects.get_or_create(exam=exam, user=user)
    exam_instance.is_extended_time = is_extended_time
    exam_instance.save()

    return JsonResponse({'is_extended_time': exam_instance.is_extended_time})

@login_required
def problem_database_view(request):
    # pass the types of categories as context
    # probabably need to make ajax call to grab which type of questions, but that's for later
    # for now, just pass the types of categories
    category_data = []
    for key in sat_math_category_map.keys():
        category_data.append({
            'key': key,
            'value': sat_math_category_map[key]
        })
    context = {
        'category_data': category_data,
    }
    return render(request, 'exams/problem_database.html', context)

@login_required
def problem_database_data_view(request):
    if request.is_ajax():
        # get the category and section type selected
        # make query and return the question data as json

        """
            TODO:
            - handle other types of questions
            - handle free response questions
        """
        data = dict(request.GET.lists())
        category = data['category'][0]
        exam_type = data['exam_type'][0]
        question_type = data['question_type'][0]
        SAT, ACT = ('SAT', 'ACT')
        MATH, GRAMMAR = ('MATH', 'GRAMMAR')
        # get questions that have this exam type and question type
        if exam_type == SAT:
            exams = Exam.objects.filter(type=SAT)
            if question_type == MATH:
                sections = Section.objects.filter(exam__in=exams, type__in=['math1', 'math2'])
            elif question_type == GRAMMAR:
                sections = Section.objects.filter(exam__in=exams, type='writing')
        elif exam_type == ACT:
            exams = Exam.objects.filter(type=ACT)
            if question_type == MATH:
                sections = Section.objects.filter(exam__in=exams, type='math')
            elif question_type == GRAMMAR:
                sections = Section.objects.filter(exam__in=exams, type='english')
        else:
            if question_type == MATH:
                sections = Section.objects.filter(type__in=['math', 'math1', 'math2'])
            elif question_type == GRAMMAR:
                sections = Section.objects.filter(type__in=['english', 'writing'])

        matching_questions = []
        questions = Question.objects.filter(section__in=sections)
        """
            If question type is GRAMMAR, then need to be able to group the questions by matching section and passage number
            use dictionary to map section/passage number to a list of questions
            
            first extract section and passage number from each question (combine into a string), then append to dictionary
            then convert dictionary into expected format
            each value in the dictionary is a list of questions grouped by passage
        """
        questions_groupby_passage = dict()
        for question in questions:
            categories = question.categories.split(',') if question.categories else []
            if category in categories:
                choices_ = question.get_answers()
                choices = []
                for choice in choices_:
                    choice_ = {
                        'letter': choice.letter,
                        'text': choice.text
                    }
                    choices.append(choice_)

                question_ = {
                    'exam': question.exam.name,
                    'section': question.section.name,
                    'questionNumber': question.question_number,
                    'text': question.text,
                    'correctAnswer': question.correct_answer,
                    'choices': choices,
                    'materialUrl': question.material.url if question.material.name else None,
                }

                if question_type == GRAMMAR:
                    material_url = getMaterialUrlFromQuestion(question)
                    section_passage_str = f"{question.section};{question.passage}"
                    if section_passage_str not in questions_groupby_passage:
                        obj = {
                            'materialUrl': material_url,	
                            'questions': [question_]
                        }
                        questions_groupby_passage[section_passage_str] = obj
                    else:
                        questions_groupby_passage[section_passage_str]['questions'].append(question_)
                    
                matching_questions.append(question_)

    questions_groupby_passage = [val for val in questions_groupby_passage.values()]
    # CURRENTLY LIMITING TO ONLY 20 QUESTIONS
    return JsonResponse({
        'data': matching_questions[0:20],
        'questions_groupby_passage': questions_groupby_passage
    })

@login_required 
def problem_database_button_data_view(request):
    if request.is_ajax():   
        """
            return the data for the buttons
            object with the following structure:
            {
                'examType',
                'questionType',
                'category',
            }
        """
        data = dict(request.GET.lists())
        question_type = data['question_type'][0]
        exam_type = data['exam_type'][0]

        SAT, ACT = ('SAT', 'ACT')
        MATH, GRAMMAR = ('MATH', 'GRAMMAR')
        button_data = []
        if question_type == MATH:
            if exam_type == SAT:
                category_map = sat_math_category_map
            elif exam_type == ACT:
                category_map = act_math_category_map
        elif question_type == GRAMMAR:
            if exam_type == SAT:
                category_map = sat_english_category_map
            elif exam_type == ACT:
                category_map = act_english_category_map

        for key in category_map.keys():
            button_data.append({
                'key': key,
                'value': category_map[key]
            })
        return JsonResponse({'data': button_data})


@login_required
# Deletes the 'Result' and 'Student_Answer' objects and redirects to the respective section directions URL
def section_reset_view(request, pk, section_name):
    exam = Exam.objects.get(pk=pk)
    section = Section.objects.get(exam=exam, type=section_name)
    user = request.user

    if Result.objects.filter(user=user, exam=exam, section=section).exists():
        Result.objects.filter(user=user, exam=exam, section=section).delete()

    if Student_Answer.objects.filter(user=user, exam=exam, section=section_name).exists():
        Student_Answer.objects.filter(user=user, exam=exam, section=section_name).delete()

    if SectionInstance.objects.filter(user=user, exam=exam, section=section).exists():
        SectionInstance.objects.filter(user=user, exam=exam, section=section).delete()

    return redirect('exams:section-directions-view', pk=pk, section_name=section_name)

@login_required
def start_exam_view(request, pk):
    exam = Exam.objects.get(pk=pk)
    context = {
        'exam':exam,
    }
    return render(request, 'exams/start_exam.html', context)

@login_required
def section_directions_view(request, pk, section_name):
    exam = Exam.objects.get(pk=pk)
    current_section = Section.objects.get(exam=exam, type=section_name)
    user = request.user

    #Check if this section has been completed before
    if Result.objects.filter(user=user, exam=exam, section=current_section).exists():
        # If completed before, look for a section that has not been completed
        sections = exam.get_sections()
        for section in sections:
            # Check if this section has been done before
            if not Result.objects.filter(user=user, exam=exam, section=section).exists():
                return redirect('exams:section-directions-view', pk=pk, section_name=section.type)

        # Reached end of for loop, thus all sections have been done, redirect to exam list view
        return redirect('exams:exam-list-view')

    # Section has not been completed before, continue as expected
    context = {
        'exam':exam,
        'section':current_section,
    }
    return render(request, 'exams/section_directions.html', context)

@login_required
def save_timer_view(request, pk, section_name):

    data = request.POST
    section = Section.objects.get(exam_id=pk, type=section_name)
    exam = Exam.objects.get(pk=pk)
    user = request.user

    # Save the timer
    if SectionInstance.objects.filter(user=user, exam=exam, section=section):
        section_instance = SectionInstance.objects.get(user=user, exam=exam, section=section)
        section_instance.minutes_left = data['minutes']
        section_instance.seconds_left = data['seconds']
        section_instance.save()

    # Set this exam as the 'recent_exam'
    student = Student.objects.get(user=user)
    student.recent_exam = exam
    student.save()
    print('RECENT EXAM: ' + student.recent_exam.name)

    return JsonResponse({})

@login_required
# DISPLAYS THE SECTION
def section_view(request, pk, section_name):
    section = Section.objects.get(exam_id=pk, type=section_name)
    exam = Exam.objects.get(pk=pk)
    questions = Question.objects.filter(exam=exam, section=section).order_by('question_number')
    user = request.user
    minutes_remaining = None
    seconds_remaining = None

    # Check if a SectionInstance exists (section has been previously started)
    if SectionInstance.objects.filter(user=user, exam=exam, section=section).exists():
        section_instance = SectionInstance.objects.get(user=user, exam=exam, section=section)
        minutes_remaining = section_instance.minutes_left
        seconds_remaining = section_instance.seconds_left
    # Create a SectionInstance object and modify the time if extended time has been selected
    else:
        # Default ExamInstance object has is_extended_time set to false
        exam_instance, created = ExamInstance.objects.get_or_create(user=user, exam=exam)
        section_minutes = (section.time * 1.5) if exam_instance.is_extended_time else section.time
        time_object = SectionInstance.objects.create(user=user, exam=exam, section=section, minutes_left=section_minutes, seconds_left=0)
        minutes_remaining = section_minutes
        seconds_remaining = 0

    context = {
        'section':section,
        'exam':exam,
        'questions':questions,
        'minutes_remaining': minutes_remaining,
        'seconds_remaining': seconds_remaining,
    }

    template = 'exams/section.html'
    return render(request, template, context)

@login_required
# DISPLAYS THE SECTION BUT ALSO PASSES IN THE CORRECT ANSWERS
def section_review_view(request, pk, section_name):
    section = Section.objects.get(exam_id=pk, type=section_name)
    exam = Exam.objects.get(pk=pk)
    questions = Question.objects.filter(exam=exam, section=section).order_by('question_number')
    user = request.user

    correct_answers = []
    # since the previously selected answers will already be handled by the section_data view,
    # we only need to worry about the correct answers that need to be displayed

    for q in questions:
        for a in q.get_answers():
            # Condition for the correct answer being the numeric free response answer
            if not q.correct_answer.isalpha():
                correct_answers.append(q.correct_answer)
                break
            elif a.letter == q.correct_answer:
                correct_answers.append(a.text)
    context = {
        'section':section,
        'exam':exam,
        'questions':questions,
        'correct_answers':correct_answers,
    }

    template = 'exams/section.html'

    return render(request, template, context)

@login_required
def section_math_data_view(request, pk, section_name):
    # CURRENTLY HANDLES RETIEVING DATA FOR THE MATH SECTION
    section = Section.objects.get(exam=pk, type=section_name)
    exam = Exam.objects.get(pk=pk)
    user = request.user

    image_urls = []

    data = []
    #gives key, value pairs to "questions," which are the questions and the answers
    for q in Question.objects.filter(exam=exam, section=section).order_by('question_number'):
        answers = []
        for a in q.get_answers():
            answers.append(a.text)

        # ERROR OCCURS WHEN UNCOMMENTED
        #if 'no question' in q.text:
        #    q.text = ''

        # Get's the previously answered question if possible
        student_answer = Student_Answer.objects.filter(user=user, exam=exam, section=section_name, question_number=q.question_number).first()
        student_answer_text = None
        is_open_ended_correct = False
        # Case where student answer is a multiple choice
        if student_answer is not None and student_answer.answer != 'N' and student_answer.answer.isalpha():
            student_answer_text = Answer.objects.get(question=q, letter=student_answer.answer).text
        # Case where student answer is free response
        elif student_answer is not None and student_answer.answer != 'N':
            student_answer_text = student_answer.answer
            if student_answer.is_correct:
                is_open_ended_correct = True


        #GETS THE IMAGE URLS
        image_url = None

        if q.material.name:
            image_url = q.material.url

        data.append({q.question_number: [str(q), answers, student_answer_text, image_url, is_open_ended_correct]})

    return JsonResponse({
        'data': data,
        'time': section.time,
        'img_urls': image_urls,
        'section': section.type,
    })

@login_required
def section_passage_data_view(request, pk, section_name, passage_num):
    section = Section.objects.get(exam=pk, type=section_name)
    exam = Exam.objects.get(pk=pk)
    user = request.user

    image_urls = []
    if section_name == 'reading' or section_name == 'writing' or section_name == 'english' or section_name == 'science':
        image = Question.objects.filter(section=section, passage=passage_num).order_by('question_number')[0].material
        # MAKES SURE THAT A URL EXISTS
        if image.name != '':
            image_urls.append(image.url)

    # if section_name == 'math1' or section_name == 'math2':
    #     for question in section.get_questions():
    #         if question.material.name:
    #             image_urls.append(question.material.url)

    data = []
    #gives key, value pairs to "questions," which are the questions and the answers

    for q in Question.objects.filter(exam=exam, section=section, passage=passage_num).order_by('question_number'):
        answers = []
        for a in q.get_answers():
            answers.append(a.text)

        # Get's the previously answered question if possible
        student_answer = Student_Answer.objects.filter(user=user, exam=exam, section=section_name, question_number=q.question_number).first()
        student_answer_text = None
        #import pdb; pdb.set_trace()
        if student_answer is not None and student_answer.answer != 'N':
            student_answer_text = Answer.objects.get(question=q, letter=student_answer.answer).text
            student_answer_text = student_answer_text.replace('"', '&quot;')
        data.append({q.question_number: [str(q), answers, student_answer_text]})

    return JsonResponse({
        'data': data,
        'time': section.time,
        'img_urls': image_urls,
        'section': section.type,
    })

@login_required
def section_break_view(request, pk, break_num, next_section_name):
    exam = Exam.objects.get(pk=pk)
    next_section = Section.objects.get(exam=exam, type=next_section_name)
    user = request.user

    #Check if next section has been completed before
    if Result.objects.filter(user=user, exam=exam, section=next_section).exists():
        # If completed before, look for a section that has not been completed
        sections = exam.get_sections()
        for section in sections:
            # Check if this section has been done before
            if not Result.objects.filter(user=user, exam=exam, section=section).exists():
                return redirect('exams:section-directions-view', pk=pk, section_name=section.type)

        # Reached end of for loop, thus all sections have been done, redirect to exam list view
        return redirect('exams:exam-list-view')


    context = {
        'exam':exam,
        'next_section':next_section,
    }
    return render(request, 'exams/break.html', context)

@login_required
# SAVES THE SECTION TO THE DATABASE, CREATES RESULT OBJECT, DELETES SectionInstance object
def save_section_view(request, pk, section_name):
    if request.is_ajax():

        #data = request.POST
        #converts the data from a QueryDict to a dict
        #data_ = dict(data.lists())
        #gets rid of the csrf token from the dict
        #data_.pop('csrfmiddlewaretoken')

        #grabs the questions and section displayed on the site
        user = request.user
        print('USER: ' + str(user))
        section = Section.objects.get(type=section_name, exam=pk)
        questions = section.get_questions()
        exam = Exam.objects.get(pk=pk)
        raw_score = 0

        # below code isn't doing anything?
        #student_answers = Student_Answer.objects.filter(user=user, exam=exam, section=section_name)
        questions = Question.objects.filter(exam=exam, section=section)

        for question in questions:
            if Student_Answer.objects.filter(user=user, exam=exam, section=section_name, question_number=question.question_number).exists():
                #'selected_answer' is the student's answer
                selected_answer_object = Student_Answer.objects.filter(user=user, exam=exam, section=section_name, question_number=question.question_number)[0]
                selected_answer = selected_answer_object.answer
                correct_answer = question.correct_answer
                # handling the multiple choice correct answers
                if correct_answer.isalpha():
                    if selected_answer == correct_answer:
                        raw_score += 1
                        selected_answer_object.is_correct = True
                        selected_answer_object.save()
                # handling the free response correct answers
                elif selected_answer != 'N' and selected_answer != '':
                    #import pdb; pdb.set_trace()
                    correct_answers = correct_answer.split(',')
                    
                    for correct_answer in correct_answers:
                        # avoid issue where users can submit 0.5 or .5, for example
                        selected_answer = str(selected_answer)
                        if selected_answer[0] == '.':
                            selected_answer = '0' + selected_answer

                        if correct_answer[0] == '.':
                            correct_answer = '0' + correct_answer

                        selected_answer = selected_answer if '/' in selected_answer else float(selected_answer)
                        correct_answer = correct_answer if '/' in correct_answer else float(correct_answer)

                        if selected_answer == correct_answer:
                            raw_score += 1
                            selected_answer_object.is_correct = True
                            selected_answer_object.save()
                            break
            else:
                Student_Answer.objects.create(user=user, exam=exam, section=section_name, question=question, question_number=question.question_number, answer='N')

        #CREATE SECTION RESULT OBJECT ONLY IF IT DOES NOT EXIST
        if not Result.objects.filter(user=user, exam=exam, section=section).exists():
            # Calculate scaled score for the section depending on the exam type
            module_dir = os.path.dirname(__file__)
            scoring_sheet_file_path = os.path.join(module_dir, 'scoring_files/Scoring.csv')
            scoring_df = pd.read_csv(scoring_sheet_file_path)
            scaled_score = None
            
            if exam.type == 'SAT' or exam.type == 'ACT':
                section_ = section.type
                section_ = 'math' if section_ == 'math1' or section_ == 'math2' else section_
                if section_ == 'math' and exam.type == 'SAT':
                    other_math_section = 'math1' if section.type == 'math2' else 'math2'
                    other_math_section = Section.objects.get(exam=exam, type=other_math_section)
                    if Result.objects.filter(user=user, exam=exam, section=other_math_section).exists():
                        other_math_section_result = Result.objects.get(user=user, exam=exam, section=other_math_section)
                        raw_score += other_math_section_result.raw_score
                        score_index = scoring_df.loc[scoring_df[f"{exam.type}_score"] == raw_score].index[0]
                        scaled_score = scoring_df.loc[score_index, f"{exam.type}_{section_}"] // 2
                        other_math_section_result.scaled_score = scaled_score 
                        other_math_section_result.save()
                else:
                    score_index = scoring_df.loc[scoring_df[f"{exam.type}_score"] == raw_score].index[0]
                    scaled_score = scoring_df.loc[score_index, f"{exam.type}_{section_}"]
            
            """
            if exam.type == 'SAT':
                if section.type == 'reading':
                    scaled_score = 400 - 5 * (52 - raw_score)
                    scaled_score = round(scaled_score/10) * 10
                    scaled_score = 100 if scaled_score < 100 else scaled_score
                elif section.type == 'writing':
                    scaled_score = 400 - 15 * (44 - raw_score)
                    scaled_score = round(scaled_score/10) * 10
                    scaled_score = 100 if scaled_score < 100 else scaled_score
                elif section.type == 'math1':
                    scaled_score = 340 - 10 * (25 - raw_score)
                    scaled_score = round(scaled_score/10) * 10
                    scaled_score = 100 if scaled_score < 100 else scaled_score
                elif section.type == 'math2':
                    scaled_score = 460 - 10 * (38 - raw_score)
                    scaled_score = round(scaled_score/10) * 10
                    scaled_score = 100 if scaled_score < 100 else scaled_score
            elif exam.type == 'ACT':
                if section.type == 'english':
                    scaled_score = 36 - 0.5 * (75 - raw_score)
                    scaled_score = round(scaled_score)
                    scaled_score = 1 if scaled_score < 1 else scaled_score
                elif section.type == 'math':
                    scaled_score = 36 - 0.6 * (60 - raw_score)
                    scaled_score = round(scaled_score)
                    scaled_score = 1 if scaled_score < 1 else scaled_score
                elif section.type == 'reading':
                    scaled_score = 36 - 0.7 * (40 - raw_score)
                    scaled_score = round(scaled_score)
                    scaled_score = 1 if scaled_score < 1 else scaled_score
                elif section.type == 'science':
                    scaled_score = 36 - (40 - raw_score)
                    scaled_score = round(scaled_score)
                    scaled_score = 1 if scaled_score < 1 else scaled_score
            """

            Result.objects.create(section=section, user=user, exam=exam, raw_score=raw_score, scaled_score=scaled_score)

        #Deletes SectionInstance object, since the section has been finished
        if SectionInstance.objects.filter(user=user, exam=exam, section=section).exists():
            SectionInstance.objects.get(user=user, exam=exam, section=section).delete()

    return JsonResponse({'section_name':section_name})

@login_required
def save_question_view(request, pk, section_name):
    print("received request")

    if request.is_ajax():

        data = request.POST
        #converts the data from a QueryDict to a dict
        data_ = dict(data.lists())

        user = request.user

        exam = Exam.objects.get(pk=pk)
        section = Section.objects.get(type=section_name, exam=pk)

        #CONVERT THE QUOTATION ESCAPES BACK TO REGULAR QUOTES
        question_text = data['question'].replace('"', '&quot;')
        question = Question.objects.get(text=question_text, section=section, exam=exam, question_number=data['question_number'])

        # Check if a the Answer object exists (implies multiple choice)
        if Answer.objects.filter(question=question, text=data['answer']).exists():
            selected_answer = Answer.objects.get(question=question, text=data['answer']).letter
        else:
            # Implies that this question is open ended math
            selected_answer = data['answer'].replace(' ', '')

        # Checks if there has already been an answer to the question and updates if true, otherwise created the student_answer object
        if Student_Answer.objects.filter(user=user, exam=exam, section=section_name, question_number=question.question_number).exists():
            student_answer = Student_Answer.objects.filter(question_number=question.question_number, section=section_name, exam=exam, user=user)[0]
            student_answer.answer = selected_answer
            student_answer.save()
        else:
            Student_Answer.objects.create(
                answer = selected_answer,
                question_number = question.question_number,
                question = question,
                section = section_name,
                exam = exam,
                user = user,
        )
        print("CREATED STUDENT ANSWER OBJECT")
    return JsonResponse({
        'hello':'hello'
    })

@login_required
def get_next_section_view(request, pk, section_name):
    exam = Exam.objects.get(pk=pk)
    section = Section.objects.get(exam=exam, type=section_name)
    user = request.user
    num_sections = exam.get_sections().count()
    section_index = section.ordering + 1
    has_completed_exam = True
    next_section = None

    for i in range(1, num_sections):
        print('SECTION INDEX: ' + str(section_index))
        # If section_index > num_sections, start searching for available section at ordering 1
        if section_index > num_sections:
            section_index = 1

        next_section = Section.objects.get(exam=exam, ordering=section_index)
        if not Result.objects.filter(user=user, exam=exam, section=next_section).exists():
            has_completed_exam = False
            break

        section_index += 1
    # Checks if we did not find an unfinished section (exam is finished)
    if has_completed_exam:
        next_section = None
    else:
        next_section = next_section.type

    context = {
        'next_section': next_section,
        'has_completed_exam': has_completed_exam,
    }

    return JsonResponse(context)
